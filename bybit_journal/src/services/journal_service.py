"""Service layer shared by the desktop bridge and the rest of the backend."""

import json
import os
import shutil
import time
import xml.etree.ElementTree as ET
from datetime import datetime, timedelta
from pathlib import Path

import config
from api import get_wallet_balance
from db import (
    delete_trade_by_id,
    get_trade_by_id,
    get_trade_stats,
    init_db,
    insert_trade,
    query_trades,
    update_trade_journal_fields,
)
from models import Trade
import requests
from sync import sync_executions_from_category

MAX_BYBIT_RANGE_DAYS = 7
ONE_DAY_MS = 24 * 60 * 60 * 1000
ECB_DAILY_RATES_URL = "https://www.ecb.europa.eu/stats/eurofxref/eurofxref-daily.xml"
SUPPORTED_DISPLAY_CURRENCIES = ("USD", "JPY", "GBP", "CHF", "EUR")
STABLECOIN_PREFIXES = ("USD",)
STABLECOIN_TICKERS = {"USDT", "USDC", "USDE", "DAI", "FDUSD", "TUSD"}
FX_CACHE_TTL_SECONDS = 6 * 60 * 60

_fx_cache = {
    "fetched_at": 0.0,
    "rates": None,
}

DEFAULT_APP_SETTINGS = {
    "auto_sync_on_startup": True,
    "default_sync_days": 7,
}


def initialize_runtime() -> None:
    """Prepare folders, .env loading and SQLite before the UI starts."""
    config.ensure_directories()
    config.ensure_env_file()
    config.load_environment()
    init_db()


def _trade_to_dict(trade) -> dict:
    """Serialize a Trade object into a UI-friendly dictionary."""
    return {
        "id": trade.id,
        "bybit_trade_id": trade.bybit_trade_id,
        "symbol": trade.symbol,
        "side": trade.side,
        "qty": trade.qty,
        "entry_price": trade.entry_price,
        "exit_price": trade.exit_price,
        "take_profit": trade.take_profit,
        "stop_loss": trade.stop_loss,
        "leverage": trade.leverage,
        "pnl": trade.pnl,
        "invested_amount": trade.invested_amount,
        "trade_time": trade.trade_time,
        "note": trade.note,
        "screenshot_path": trade.screenshot_path,
    }


def _coerce_sync_days(value) -> int:
    """Normalize sync day input into a safe integer value."""
    try:
        parsed = int(value)
    except (TypeError, ValueError):
        parsed = DEFAULT_APP_SETTINGS["default_sync_days"]
    return max(parsed, 1)


def load_app_settings() -> dict:
    """Load persisted app settings with defaults."""
    config.ensure_directories()
    if not config.SETTINGS_PATH.exists():
        return DEFAULT_APP_SETTINGS.copy()

    try:
        payload = json.loads(config.SETTINGS_PATH.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return DEFAULT_APP_SETTINGS.copy()

    return {
        "auto_sync_on_startup": bool(payload.get("auto_sync_on_startup", True)),
        "default_sync_days": _coerce_sync_days(payload.get("default_sync_days")),
    }


def save_app_settings(auto_sync_on_startup: bool, default_sync_days: int) -> dict:
    """Persist app settings used by the desktop shell."""
    config.ensure_directories()
    payload = {
        "auto_sync_on_startup": bool(auto_sync_on_startup),
        "default_sync_days": _coerce_sync_days(default_sync_days),
    }
    config.SETTINGS_PATH.write_text(json.dumps(payload, indent=2), encoding="utf-8")
    return payload


def get_api_status_data() -> dict:
    """Return the current API configuration state for the desktop UI."""
    has_credentials = config.has_api_credentials()
    runtime_paths = config.get_runtime_paths()
    return {
        "has_credentials": has_credentials,
        "dev_mode": config.DEV_MODE,
        "settings": load_app_settings(),
        **runtime_paths,
        "message": (
            "API configured"
            if has_credentials
            else "API keys missing"
        ),
    }


def get_trades_data(
    symbol: str | None = None,
    side: str | None = None,
    start_time: str | None = None,
    end_time: str | None = None,
    min_pnl: float | None = None,
    max_pnl: float | None = None,
    limit: int | None = None,
) -> dict:
    """Return serialized trades together with the effective filters."""
    trades = query_trades(
        symbol=symbol,
        side=side,
        start_time=start_time,
        end_time=end_time,
        min_pnl=min_pnl,
        max_pnl=max_pnl,
        limit=limit,
    )

    return {
        "count": len(trades),
        "filters": {
            "symbol": symbol,
            "side": side,
            "start_time": start_time,
            "end_time": end_time,
            "min_pnl": min_pnl,
            "max_pnl": max_pnl,
            "limit": limit,
        },
        "trades": [_trade_to_dict(trade) for trade in trades],
    }


def _build_excel_workbook(trades: list[dict], stats: dict, exported_at: str):
    """Build the Excel workbook used by the export action."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError as error:
        raise RuntimeError(
            "Excel export unavailable: install openpyxl in the project venv."
        ) from error

    workbook = Workbook()
    trades_sheet = workbook.active
    trades_sheet.title = "Trades"

    header_fill = PatternFill(fill_type="solid", fgColor="2A174B")
    header_font = Font(color="FFFFFF", bold=True)
    title_font = Font(size=14, bold=True)
    centered = Alignment(vertical="center")

    trades_sheet["A1"] = "Bybit Trade Journal - Export Excel"
    trades_sheet["A1"].font = title_font
    trades_sheet["A2"] = f"Genere le {exported_at}"

    headers = [
        "Trade ID",
        "Symbol",
        "Side",
        "Qty",
        "Entry Price",
        "Exit Price",
        "PnL",
        "Invested Amount",
        "Trade Time",
        "Note",
    ]

    for column_index, header in enumerate(headers, start=1):
        cell = trades_sheet.cell(row=4, column=column_index, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = centered

    for row_index, trade in enumerate(trades, start=5):
        trades_sheet.cell(row=row_index, column=1, value=trade.get("bybit_trade_id"))
        trades_sheet.cell(row=row_index, column=2, value=trade.get("symbol"))
        trades_sheet.cell(row=row_index, column=3, value=trade.get("side"))
        trades_sheet.cell(row=row_index, column=4, value=trade.get("qty"))
        trades_sheet.cell(row=row_index, column=5, value=trade.get("entry_price"))
        trades_sheet.cell(row=row_index, column=6, value=trade.get("exit_price"))
        trades_sheet.cell(row=row_index, column=7, value=trade.get("pnl"))
        trades_sheet.cell(row=row_index, column=8, value=trade.get("invested_amount"))
        trades_sheet.cell(row=row_index, column=9, value=trade.get("trade_time"))
        trades_sheet.cell(row=row_index, column=10, value=trade.get("note"))

    trades_sheet.freeze_panes = "A5"
    trades_sheet.column_dimensions["A"].width = 18
    trades_sheet.column_dimensions["B"].width = 14
    trades_sheet.column_dimensions["C"].width = 10
    trades_sheet.column_dimensions["D"].width = 10
    trades_sheet.column_dimensions["E"].width = 14
    trades_sheet.column_dimensions["F"].width = 14
    trades_sheet.column_dimensions["G"].width = 12
    trades_sheet.column_dimensions["H"].width = 16
    trades_sheet.column_dimensions["I"].width = 20
    trades_sheet.column_dimensions["J"].width = 24

    stats_sheet = workbook.create_sheet(title="Stats")
    stats_sheet["A1"] = "Synthese"
    stats_sheet["A1"].font = title_font
    stats_sheet["A2"] = f"Genere le {exported_at}"

    stats_rows = [
        ("Total trades", stats.get("total_trades", 0)),
        ("Total PnL", stats.get("total_pnl", 0)),
        ("Average PnL", stats.get("average_pnl", 0)),
        ("Winning trades", stats.get("winning_trades", 0)),
        ("Losing trades", stats.get("losing_trades", 0)),
        ("Breakeven trades", stats.get("breakeven_trades", 0)),
        ("Win rate", stats.get("win_rate", 0)),
        ("Best trade", stats.get("best_trade", 0)),
        ("Worst trade", stats.get("worst_trade", 0)),
        ("Invested amount", stats.get("total_invested_amount", 0)),
    ]

    for row_index, (label, value) in enumerate(stats_rows, start=4):
        label_cell = stats_sheet.cell(row=row_index, column=1, value=label)
        value_cell = stats_sheet.cell(row=row_index, column=2, value=value)
        label_cell.fill = header_fill
        label_cell.font = header_font
        label_cell.alignment = centered
        value_cell.alignment = centered

    stats_sheet.column_dimensions["A"].width = 24
    stats_sheet.column_dimensions["B"].width = 18

    return workbook


def get_trade_stats_data(
    symbol: str | None = None,
    side: str | None = None,
    start_time: str | None = None,
    end_time: str | None = None,
) -> dict:
    """Return aggregate stats plus the filters used to compute them."""
    stats = get_trade_stats(
        symbol=symbol,
        side=side,
        start_time=start_time,
        end_time=end_time,
    )
    recent_trades = query_trades(
        symbol=symbol,
        side=side,
        start_time=start_time,
        end_time=end_time,
        limit=12,
    )
    stats["filters"] = {
        "symbol": symbol,
        "side": side,
        "start_time": start_time,
        "end_time": end_time,
    }
    stats["chart_points"] = [
        {
            "label": (trade.trade_time or "")[5:16] or trade.symbol,
            "pnl": trade.pnl,
            "symbol": trade.symbol,
        }
        for trade in reversed(recent_trades)
    ]
    return stats


def get_dashboard_data() -> dict:
    """Assemble the dashboard payload displayed by the frontend home view."""
    recent_trades = get_trades_data(limit=5)
    stats = get_trade_stats_data()
    api_status = get_api_status_data()
    wallet = get_wallet_summary() if api_status["has_credentials"] else {
        "success": False,
        "error": "API configuration required.",
        "account": None,
        "non_zero_balances": [],
    }

    return {
        "api_status": api_status,
        "stats": stats,
        "wallet": wallet,
        "recent_trades": recent_trades["trades"],
        "recent_trade_count": recent_trades["count"],
    }


def get_trade_detail_data(trade_id: int) -> dict:
    """Return one full trade payload for the detail view."""
    trade = get_trade_by_id(trade_id)
    if not trade:
        raise ValueError(f"Trade not found: id={trade_id}")

    payload = _trade_to_dict(trade)
    payload["has_note"] = bool((trade.note or "").strip())
    payload["has_screenshot_path"] = bool((trade.screenshot_path or "").strip())
    return payload


def update_trade_journal_data(trade_id: int, note: str | None, screenshot_path: str | None) -> dict:
    """Save the editable journal fields for a trade."""
    updated = update_trade_journal_fields(
        trade_id=trade_id,
        note=(note or "").strip() or None,
        screenshot_path=(screenshot_path or "").strip() or None,
    )
    if not updated:
        raise ValueError(f"Trade not found: id={trade_id}")

    return {
        "updated": True,
        "trade": get_trade_detail_data(trade_id),
    }


def export_trades_to_excel(
    symbol: str | None = None,
    side: str | None = None,
    start_time: str | None = None,
    end_time: str | None = None,
    min_pnl: float | None = None,
    max_pnl: float | None = None,
    limit: int | None = None,
) -> dict:
    """Export filtered trades and summary stats to an Excel workbook."""
    export_data = get_trades_data(
        symbol=symbol,
        side=side,
        start_time=start_time,
        end_time=end_time,
        min_pnl=min_pnl,
        max_pnl=max_pnl,
        limit=limit,
    )
    stats = get_trade_stats_data(
        symbol=symbol,
        side=side,
        start_time=start_time,
        end_time=end_time,
    )
    exported_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    workbook = _build_excel_workbook(export_data["trades"], stats, exported_at)

    config.ensure_directories()
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    export_path = config.EXPORTS_DIR / f"bybit_trades_{timestamp}.xlsx"
    workbook.save(export_path)

    return {
        "exported": True,
        "path": str(export_path),
        "count": export_data["count"],
        "filters": export_data["filters"],
        "generated_at": exported_at,
    }


def delete_trade_data(trade_id: int) -> dict:
    """Delete one stored trade and return a small UI-friendly summary."""
    deleted = delete_trade_by_id(trade_id)
    if not deleted:
        raise ValueError(f"Trade not found: id={trade_id}")

    return {
        "deleted": True,
        "trade_id": trade_id,
    }


def open_runtime_folder(target: str) -> dict:
    """Open one user-facing runtime folder in the system file explorer."""
    runtime_paths = config.get_runtime_paths()
    folder_map = {
        "config": Path(runtime_paths["config_dir"]),
        "data": Path(runtime_paths["data_dir"]),
        "exports": Path(runtime_paths["exports_dir"]),
        "backups": Path(runtime_paths["backups_dir"]),
    }
    if target not in folder_map:
        raise ValueError(f"Unknown folder target: {target}")

    path = folder_map[target]
    path.mkdir(parents=True, exist_ok=True)

    if os.name == "nt":
        os.startfile(path)  # type: ignore[attr-defined]
    else:
        raise RuntimeError("Folder opening is only supported on Windows.")

    return {
        "opened": True,
        "target": target,
        "path": str(path),
    }


def create_database_backup() -> dict:
    """Create a timestamped copy of the SQLite database."""
    config.ensure_directories()
    if not config.DB_PATH.exists():
        raise RuntimeError("Database file not found.")

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_path = config.BACKUPS_DIR / f"journal_backup_{timestamp}.db"
    shutil.copy2(config.DB_PATH, backup_path)
    return {
        "created": True,
        "path": str(backup_path),
        "filename": backup_path.name,
    }


def restore_database_backup(backup_path: str) -> dict:
    """Restore the active database from a provided backup file."""
    source = Path(backup_path).expanduser()
    if not source.exists():
        raise ValueError("Backup file not found.")
    if source.suffix.lower() not in {".db", ".sqlite", ".sqlite3"}:
        raise ValueError("Unsupported backup file.")

    config.ensure_directories()
    shutil.copy2(source, config.DB_PATH)
    init_db()
    return {
        "restored": True,
        "path": str(source),
        "db_path": str(config.DB_PATH),
    }


def get_wallet_summary(wallet: dict | None = None) -> dict:
    """Normalize the Bybit wallet payload for the dashboard wallet widgets."""
    wallet_data = wallet if wallet is not None else get_wallet_balance()

    if wallet_data.get("retCode") != 0:
        return {
            "success": False,
            "error": wallet_data.get("retMsg") or "Unknown Bybit error.",
            "account": None,
            "non_zero_balances": [],
        }

    accounts = wallet_data.get("result", {}).get("list", [])
    if not accounts:
        return {
            "success": False,
            "error": "No account returned by the API.",
            "account": None,
            "non_zero_balances": [],
        }

    account = accounts[0]
    non_zero_balances = []
    stable_balances = []
    for coin in account.get("coin", []):
        equity = float(coin.get("equity", 0) or 0)
        if equity > 0:
            balance = {
                "coin": coin.get("coin"),
                "equity": coin.get("equity"),
                "wallet_balance": coin.get("walletBalance"),
                "usd_value": coin.get("usdValue"),
            }
            non_zero_balances.append(balance)
            coin_name = str(balance["coin"] or "").upper()
            if coin_name in STABLECOIN_TICKERS or coin_name.startswith(STABLECOIN_PREFIXES):
                stable_balances.append(balance)

    non_zero_balances.sort(key=lambda item: float(item.get("usd_value", 0) or 0), reverse=True)
    stable_balances.sort(key=lambda item: float(item.get("usd_value", 0) or 0), reverse=True)
    stable_total_usd = sum(float(item.get("usd_value", 0) or 0) for item in stable_balances)

    wallet_summary = {
        "success": True,
        "error": None,
        "account": {
            "account_type": account.get("accountType"),
            "total_equity": account.get("totalEquity"),
            "total_wallet_balance": account.get("totalWalletBalance"),
        },
        "non_zero_balances": non_zero_balances,
        "stable_balances": stable_balances,
        "stable_total_usd": stable_total_usd,
    }
    wallet_summary["currency_options"] = list(SUPPORTED_DISPLAY_CURRENCIES)
    wallet_summary["conversion_rates"] = _get_usd_conversion_rates()
    wallet_summary["display_currency"] = "USD"
    return wallet_summary


def _get_usd_conversion_rates() -> dict:
    """Fetch display-currency conversion rates based on ECB daily FX data."""
    if _fx_cache["rates"] and (time.time() - _fx_cache["fetched_at"]) < FX_CACHE_TTL_SECONDS:
        return _fx_cache["rates"]

    rates_per_eur = {"EUR": 1.0}

    try:
        response = requests.get(ECB_DAILY_RATES_URL, timeout=10)
        response.raise_for_status()
        root = ET.fromstring(response.text)
        for cube in root.iter():
            currency = cube.attrib.get("currency")
            rate = cube.attrib.get("rate")
            if currency and rate:
                rates_per_eur[currency] = float(rate)
    except (requests.RequestException, ET.ParseError, ValueError):
        if _fx_cache["rates"]:
            return _fx_cache["rates"]
        return {
            "USD": 1.0,
            "JPY": 0.0,
            "GBP": 0.0,
            "CHF": 0.0,
            "EUR": 0.0,
        }

    usd_per_eur = rates_per_eur.get("USD")
    if not usd_per_eur:
        if _fx_cache["rates"]:
            return _fx_cache["rates"]
        return {
            "USD": 1.0,
            "JPY": 0.0,
            "GBP": 0.0,
            "CHF": 0.0,
            "EUR": 0.0,
        }

    usd_rates = {"USD": 1.0}
    for currency in SUPPORTED_DISPLAY_CURRENCIES:
        if currency == "USD":
            continue
        if currency == "EUR":
            usd_rates[currency] = 1 / usd_per_eur
            continue

        target_per_eur = rates_per_eur.get(currency)
        usd_rates[currency] = (target_per_eur / usd_per_eur) if target_per_eur else 0.0

    _fx_cache["fetched_at"] = time.time()
    _fx_cache["rates"] = usd_rates
    return usd_rates


def _build_sync_ranges(days: int, now_ms: int | None = None) -> list[tuple[int, int]]:
    """Split a sync request into Bybit-compatible 7-day windows."""
    safe_days = max(days, 1)
    end_ms = now_ms if now_ms is not None else int(time.time() * 1000)
    start_ms = end_ms - (safe_days * ONE_DAY_MS)
    chunk_ms = MAX_BYBIT_RANGE_DAYS * ONE_DAY_MS

    ranges = []
    current_start = start_ms
    while current_start < end_ms:
        current_end = min(current_start + chunk_ms, end_ms)
        ranges.append((current_start, current_end))
        current_start = current_end

    return ranges


def sync_bybit_trades_data(days: int = 30, now_ms: int | None = None) -> dict:
    """Synchronize Bybit executions category by category and summarize the result."""
    effective_days = _coerce_sync_days(days)
    sync_ranges = _build_sync_ranges(days=effective_days, now_ms=now_ms)
    first_start, last_end = sync_ranges[0][0], sync_ranges[-1][1]

    per_category = []
    total_inserted = 0

    for category in ["linear", "spot", "inverse", "option"]:
        inserted_count = 0
        for start_ms, end_ms in sync_ranges:
            try:
                inserted_count += sync_executions_from_category(
                    category=category,
                    start_time=start_ms,
                    end_time=end_ms,
                    limit=100,
                )
            except ValueError:
                # A category failure should not block the whole desktop refresh.
                continue
        total_inserted += inserted_count
        per_category.append(
            {
                "category": category,
                "inserted_count": inserted_count,
            }
        )

    return {
        "days": effective_days,
        "start_time_ms": first_start,
        "end_time_ms": last_end,
        "range_count": len(sync_ranges),
        "categories": per_category,
        "total_inserted": total_inserted,
    }


def configure_api_credentials(api_key: str, api_secret: str) -> dict:
    """Persist Bybit credentials into the root .env file."""
    config.ensure_env_file()

    content = (
        f"BYBIT_API_KEY={api_key.strip()}\n"
        f"BYBIT_API_SECRET={api_secret.strip()}\n"
    )
    Path(config.ENV_PATH).write_text(content, encoding="utf-8")
    config.load_environment()

    return {
        "saved": True,
        "env_path": str(config.ENV_PATH),
        "has_credentials": config.has_api_credentials(),
    }


def seed_dev_test_trades(count: int = 20) -> dict:
    """Insert synthetic trades used to test the GUI without live API data."""
    symbols = [
        "BTCUSDT",
        "ETHUSDT",
        "SOLUSDT",
        "XRPUSDT",
        "ADAUSDT",
        "AVAXUSDT",
        "LINKUSDT",
        "DOGEUSDT",
        "BNBUSDT",
        "SUIUSDT",
    ]
    inserted = 0
    base_time = datetime(2026, 3, 1, 9, 0, 0)

    for index in range(count):
        entry_price = 100 + (index * 17.5)
        qty = round(0.1 + (index * 0.03), 4)
        pnl = round(((index % 7) - 3) * 4.25, 2)
        trade_time = (base_time + timedelta(hours=index * 6)).strftime("%Y-%m-%d %H:%M:%S")
        side = "Buy" if index % 2 == 0 else "Sell"
        symbol = symbols[index % len(symbols)]

        trade = Trade(
            bybit_trade_id=f"DEVTEST-{index + 1:03d}",
            symbol=symbol,
            side=side,
            qty=qty,
            entry_price=round(entry_price, 2),
            exit_price=round(entry_price + (pnl / max(qty, 0.0001)), 2),
            take_profit=round(entry_price * 1.03, 2),
            stop_loss=round(entry_price * 0.97, 2),
            leverage=float((index % 5) + 1),
            pnl=pnl,
            invested_amount=round(entry_price * qty, 2),
            trade_time=trade_time,
            note=f"DEV TEST ONLY #{index + 1}",
            screenshot_path=None,
        )
        inserted += 1 if insert_trade(trade) else 0

    return {
        "requested": count,
        "inserted": inserted,
        "message": f"{inserted} dev trade(s) inserted.",
    }
